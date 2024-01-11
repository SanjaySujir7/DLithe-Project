
from flask import Flask, jsonify,render_template,request, send_file,redirect,flash,session
import mysql.connector
import csv
from Process import Inst_Process,Random_Password,Certificat_Number_Generator
from External import Pdf_Certificate
from datetime import datetime,timedelta
from Sideoper import Hash_Password,Clean_Data
from Sideoper import Mysql_Credentials
from Email import Certificate_Email
import Credentials
from openpyxl import load_workbook
from functools import wraps
import jwt



app = Flask(__name__)
app.secret_key = "$2b$12$VraDT1QPplopUfEiM9Gkgz2t7yXhrlK28apuN1o6ILQHEZC9yM12"



def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.cookies.get('sat')

        if not token:
            session['redirect_url'] = request.url.replace(request.url_root,'/')
    
            return redirect('/admin-login')

        try:
            data = jwt.decode(token, app.config['SECRET_KEY'],algorithms=['HS256'])

        except jwt.ExpiredSignatureError:
            session['redirect_url'] = request.endpoint
            Res = redirect('/admin-login')
            Res.delete_cookie('sat')
            return Res 
        
        except jwt.InvalidTokenError:
            session['redirect_url'] = request.endpoint
            return redirect('/admin-login')
        
        if not data['user'] == request.headers.get('X-Real-IP'):
        
            Res = redirect('/admin-login')
            Res.delete_cookie('sat')
            return Res 
            
        return f(data)

    return decorated

    
    
@app.route("/admin-certificate-email-send",methods=['POST'])
def Admin_Certificate_Email_Send (data):
    
    cred = request.get_json()
        
    if not cred['pass'] == "*tggwhw$gwg@#(0hjwjwjwj??53773&**(#$#":
        return jsonify({"res" : False})
    
    
    
    Mydb = mysql.connector.connect(
            host = "localhost",
            user = Mysql_Credentials.USER,
            password = Mysql_Credentials.PASS,
            database = 'sis'
            
            )
            
    cursor = Mydb.cursor()
    
    cursor.execute("""SELECT First_Name,Last_Name,Register_Number,Institution_Name,Start_Date,End_Date,Certificate_Number,Course_Name,
                   Email FROM students WHERE Batch = %s""",(cred['data'],))
    
    Data = cursor.fetchall()
    
    cursor.close()
    Mydb.close()
    
    if Data:
        
        if not Data[0][6]:
            return jsonify({'res' : False})
        
        Certificate_Email(Data).Send()
        
    else:
        return jsonify({"res" : False})
    

    return jsonify({'res':True})


@app.route("/students-certicate-verify-id",methods=['POST'])
def Students_Certificate_verify_id ():
    Data = request.get_json()
    
    Mydb = mysql.connector.connect(
            host = "localhost",
            user = Mysql_Credentials.USER,
            password = Mysql_Credentials.PASS,
            database = 'sis'
            
            )
            
    cursor = Mydb.cursor()
    
    cursor.execute("SELECT First_Name , Last_Name , Register_Number, Institution_Name, Start_Date , End_Date , Certificate_Number, Course_Name FROM students  WHERE Certificate_Number = %s",(Data['data'],))
    data = cursor.fetchall()
    
    if data:
        Name = data[0][0] + " " + data[0][1]
        Register = data[0][2]
        Collage = data[0][3]
        Start = data[0][4]
        End = data[0][5]
        Certificate_Id = data[0][6]
        Course_Name = data[0][7]
        
        Pdf_Certificate(Name,Register,Collage,Start,End,Certificate_Id,Course_Name).Print()
        
        return send_file('output.pdf',as_attachment=True)
    
    else:
        return jsonify({'res' : False}),401


@app.route('/students-certificate-download')
def Student_Certificate_Download():
    return render_template("student_confirm_certificate.html")


@app.route('/update-student-data',methods=['POST'])
@token_required
def Student_Data_Update (admin):
    data = request.get_json()
    
    if data:
    
        Name = data['First']
        Last = data['Last']
        Phone = data['Phone']
        Email = data['Email']
        Reg = data['Reg']
        Inst = data['Inst']
        Course = data['Course']
        Total = data['Total']
        Mode = data['Mode']
        Payment = data['Payment']

        Mydb = mysql.connector.connect(
                host = "localhost",
                user = Mysql_Credentials.USER,
                password = Mysql_Credentials.PASS,
                database = 'sis'
                
                )
                
        cursor = Mydb.cursor()
        
        cursor.execute("""UPDATE students SET Phone = %s,Email = %s,Institution_Name =%s,Course_Name =%s,Total =%s,Mode =%s,Payment_Status =%s WHERE
        First_Name = %s AND Last_Name = %s AND Register_Number = %s """,(Phone,Email,Inst,Course,Total,Mode,Payment,Name,Last,Reg,))
        
        Mydb.commit()
        cursor.close()
        Mydb.close()
        
        return jsonify({"res" : True})
        
    else:
        return jsonify({'res' : False})
        


@app.route('/bulk-action',methods= ['POST'])
@token_required
def Admin_Bulk_Action (admin):
    
    data = request.get_json()
    
    if data['for'] == "End_Date":

        Batch = data['data']['Batch']
        Date = data['data']['Date']
        
        if Batch and Date and not 'All' in Batch:
            
            Mydb = mysql.connector.connect(
            host = "localhost",
            user = Mysql_Credentials.USER,
            password = Mysql_Credentials.PASS,
            database = 'sis'
            
            )
            
            cursor = Mydb.cursor()
            
            cursor.execute("UPDATE students SET End_Date = %s WHERE Batch = %s",(Date,Batch,))
            
            Mydb.commit()
            
            cursor.close()
            Mydb.close()
            
            return jsonify({'res' : True})
        
        else:
            return jsonify({'res' : False})
        
        
    elif data['for'] == "Start_Date":
        Batch = data['data']['Batch']
        Date = data['data']['Date']
        
        if Batch and Date and not 'All' in Batch:
            
            Mydb = mysql.connector.connect(
            host = "localhost",
            user = Mysql_Credentials.USER,
            password = Mysql_Credentials.PASS,
            database = 'sis'
            
            )
            
            cursor = Mydb.cursor()
            
            cursor.execute("UPDATE students SET Start_Date = %s  WHERE  Batch = %s",(Date,Batch,))
            
            Mydb.commit()
            
            cursor.close()
            Mydb.close()
            
            return jsonify({'res' : True})
        
        else:
            return jsonify({'res' : False})
        
    else:
        return jsonify({'res' : False})
    
    

@app.route('/admin-certificate-generate-id',methods=['POST'])
@token_required
def Admin_Certificate_Generate_Id (admin):
    Data = request.get_json()
    
    
    if Data['data']:
        
        Mydb = mysql.connector.connect(
            host = "localhost",
            user = Mysql_Credentials.USER,
            password = Mysql_Credentials.PASS,
            database = 'sis'
            
            )
            
        cursor = Mydb.cursor()
        
        
        for students in Data['data']:
            First = students['First_Name']
            Phone = students['Phone']
            Email = students['Email']
            Usn = students['Usn']
            Course_Name = students['Course_Name']
            
            cursor.execute("SELECT End_Date,Certificate_Number FROM students WHERE First_Name =%s AND Phone = %s AND Email = %s AND Register_Number = %s AND Course_Name = %s",
                           (First,Phone,Email,Usn,Course_Name,))
            
            data = cursor.fetchall()
            End_Date = data[0][0]
            Certificate_Id = data[0][1]
            
            if not End_Date:
                cursor.close()
                Mydb.close()
                return jsonify({"res" : False}),406
            
            if Certificate_Id:
                pass
                
            else:
                Certificate_Id = Certificat_Number_Generator(Course_Name,End_Date).Generate()
                
                cursor.execute("UPDATE students SET Certificate_Number = %s WHERE First_Name = %s AND Phone = %s AND Email = %s AND Register_Number = %s AND Course_Name = %s", (Certificate_Id, First, Phone, Email, Usn, Course_Name))
                
                
        Mydb.commit()
        cursor.close()
        Mydb.close()
                
            
        return jsonify({'res' : True})
    
    else:
        return jsonify({'res' : False})
            
        
    

@app.route('/admin-certificate-fetch-data',methods =['POST'])
@token_required
def Admin_Certificate_Fetch_Data (admin):
    
    data = request.get_json()
    
    
    if data['pass'] == "!#@1234q:{)324++@9926xcvbn":
        
        Mydb = mysql.connector.connect(
        host = "localhost",
        user = Mysql_Credentials.USER,
        password = Mysql_Credentials.PASS,
        database = 'sis'
        
        )
        
        cursor = Mydb.cursor()
        Send_List = []
        value = data['value']
        Batch = data['batch']
        
        
        if value == "generat-selec":
            
            cursor.execute("SELECT First_Name, Last_Name, Phone, Email, Certificate_Number,Register_Number,Course_Name  FROM students WHERE Certificate_Number IS NOT NULL AND Batch = %s;",(Batch,))
            data = cursor.fetchall()
            if data:
                for cred in data:
                    send_data = {
                        'First_Name' : cred[0],
                        'Last_Name' : cred[1],
                        'Phone' : cred[2],
                        'Email' : cred[3],
                        'Certi_Number' : cred[4],
                        'Usn' : cred[5],
                        'Certi_Status' : "True",
                        'Course_Name' : cred[6]
                    }
                    
                    Send_List.append(send_data)
                    
                    cursor.close()
                    Mydb.close()
            
                return jsonify({'exists' : True , 'data' : Send_List})
                
            else:
                cursor.close()
                Mydb.close()
                return jsonify({'exists' : False})
            
            
        elif value == "nongener-selec":
            
            cursor.execute("SELECT First_Name, Last_Name, Phone, Email, Certificate_Number, Register_Number,Course_Name  FROM students WHERE Certificate_Number is NULL AND Batch = %s;",(Batch,))
            data = cursor.fetchall()
    
            if data:
                for cred in data:
                    send_data = {
                        'First_Name' : cred[0],
                        'Last_Name' : cred[1],
                        'Phone' : cred[2],
                        'Email' : cred[3],
                        'Certi_Number' : cred[4],
                        'Usn' : cred[5],
                        'Certi_Status' : "False",
                        'Course_Name' : cred[6]
                    }
                    
                    Send_List.append(send_data)
                    
                cursor.close()
                Mydb.close()
                return jsonify({'exists' : True , 'data' : Send_List})
                
            else:
                cursor.close()
                Mydb.close()
                return jsonify({'exists' : False})
            
        elif value == "errror-selec":
            
                return jsonify({'exists' : False})
            
        else:
            cursor.close()
            Mydb.close()
            return jsonify({'exists' : False})
            
    else:
        return jsonify({'error' : "access denied!"})
        


@app.route('/admin-certificate')
@token_required
def Admin_Certificate_Page (data):
    
    return render_template("Admin_Certificate_Page.html")
    
    

@app.route('/Student-Account-check',methods=["POST"])
@token_required
def Check_Account_Exists (admin):
    data = request.get_json()
    
    if data:
        Phone = data["Phone"]
        Email = data['Email']
        Name = data['Name'].lower()
        Last = data['Last'].lower()
        
        mydb = mysql.connector.connect(
                host = 'localhost',
                user = Mysql_Credentials.USER,
                password = Mysql_Credentials.PASS,
                database = 'sis'
            )
            
        cursor = mydb.cursor()
        
        cursor.execute("SELECT First_Name FROM students WHERE Phone = %s AND Email = %s AND First_Name = %s AND Last_Name = %s;",(Phone,Email,Name,Last))
        data = cursor.fetchall()
        
        cursor.close()
        mydb.close()
        
        if data:
            return jsonify({"res" : True})
        
        else:
            return jsonify({"res" : False})
    
    else:
        return jsonify({"res" : False})
    


@app.route('/certificate-verify',methods=['POST'])
def Verify_Certificate ():
    Data = request.get_json()
    
    Certificate_Id = Data['data']
    
    if Certificate_Id:
        mydb = mysql.connector.connect(
            host = 'localhost',
            user = Mysql_Credentials.USER,
            password = Mysql_Credentials.PASS,
            database = 'sis'
        )
        
        cursor = mydb.cursor()
        
        cursor.execute("SELECT  First_Name,Email FROM students WHERE Certificate_Number = %s",
                        (Certificate_Id,))
        
        data = cursor.fetchall()
        
        cursor.close()
        mydb.close()
        
        if data:
            
            return jsonify({'result':True,'title' : "Certificate is Authorized"})
        
        else:
            return jsonify({'result':False,'title' : "Certificate is Unauthorized"})
        
    else:
        return jsonify({'result':False,'title' : "Invalid Input!"})


@app.route('/')
def Landing_Page ():
    
    return render_template('Landing_Page.html')




@app.route('/add-student',methods = ['POST'])
@token_required
def Add_Student (admin):
    data = request.get_json()
    
    if data :
    
        Name = data['Name'].lower()
        Last = data['Last'].lower()
        Phone = data['Phone']
        Email = data['Email']
        Register_Number= data['Reg'].upper()
        Institution_Name = data['Inst']
        Course_Name = data['Course'].lower()
        Total = data['Total']
        Payment_Status = data['Payment'].lower()
        Mode = data['Mode']
        Entry_Date = data['Entry_Date']

        Final = True
        
        if Name and Last and Phone and Email and Register_Number and Institution_Name and Course_Name and Total and Payment_Status and Mode:
            if "@" in Email:
                if not ".com" in Email and not ".in" in  Email:
                    Final = False
                
            if not len(Phone) >= 10 :
                Final = False
                
        if Final:
            
            
            Mydb = mysql.connector.connect(
            host = "localhost",
            user = Mysql_Credentials.USER,
            password = Mysql_Credentials.PASS,
            database = 'sis'
            
            )
            
            cursor = Mydb.cursor()
            
            cursor.execute('SELECT * FROM Key_Dictionary')
            Key_List = cursor.fetchall()
            
            Key_Process = Inst_Process(Register_Number,Institution_Name,Key_List).Process()
                    
            Inst_Key = Key_Process['inst_key']
                    
            if not Key_Process['got']:
                 cursor.execute("INSERT INTO Key_Dictionary (Reg_key, Inst) VALUES (%s , %s)",Key_Process['keys'])
                        
            
            cursor.execute("SELECT Entry_Date , Inst_Key FROM students WHERE Phone = %s AND Register_Number = %s AND Course_Name = %s;",(Phone,Register_Number,Course_Name,))
            if_data_exist = cursor.fetchall()
        
            if if_data_exist :
                
                return jsonify({'res' : True,'date' : if_data_exist[0][0],'Inst' : if_data_exist[0][1]})
            
            else:
                Batch = Credentials.Batch
                Password = Random_Password(10).Generate()
                Payment_date = datetime.now()
                
                cursor.execute("""INSERT INTO students (First_Name, Last_Name, Phone,
                            Email , Register_Number, Institution_Name, Mode,Course_Name,
                            Total, Entry_Date,Payment_Status,Inst_Key,Password,Batch,Payment_Date) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);""",(Name,Last,Phone,Email,Register_Number,Institution_Name,
                            Mode,Course_Name,Total,Entry_Date,Payment_Status,Inst_Key,Password,Batch,Payment_date))
            
                
                Mydb.commit()
                
                cursor.execute('SELECT Entry_Date, Inst_Key FROM students WHERE Phone = %s AND Register_Number = %s ;',(Phone,Register_Number,))
                data = cursor.fetchall()
                
                cursor.close()
                Mydb.close()
                
            
                return jsonify({'res' : True,'date' : data[0][0],'Inst' : data[0][1]})
        
        else:
            
            return jsonify({'res' : False})
        
    else:
        return jsonify({'res' : False})




@app.route('/export-list',methods = ['POST'])
@token_required
def Export_Data (data):
    
    Data = request.get_json()
    
    if 'pass' in Data:
        
        if not Data['pass'] == "$2b$12$MAguwgtdDIGg3JAC9xKlcQC8EEinQsAfhtfs2Z7I8DAp9aG":
            return "Unauthorized!"
        
    else:
        return "Unauthorized!"


    Export_List = Data['Export_List']
    Export_Limit = Data['Export_Limit']
    Export_Format = Data['Export_Format']

    Headings = ['First_Name', 'Last_Name', 'Phone', 'Email','Register_Number', 'Institution_Name','Mode','Course_Name','Total','Entry_Date',
                'Payment_Status','Inst_Key','Password','Certificate','Start_Date','End_Date','Payment_Date','Department','Batch']
    
    Export_Headings = ['First Name', 'Last Name', 'Phone', 'Email','Usn', 'College','Mode','Domain','Total','Entry Date',
                'Payment Status','Inst Key','Password','Certificate ID','Start Date','End Date','Payment Date','Branch','Batch']
    
    New_Heading = []
    New_Export_Heading = []
    New_Export_List = []
    
    for i in range(19):
        if i in Export_Limit:
            New_Heading.append(Headings[i])
            New_Export_Heading.append(Export_Headings[i])
    
            
    for all in Export_List:
        temp_list = []
        
        for all_heading in New_Heading:
            temp_list.append(all[all_heading])
        
        New_Export_List.append(temp_list)
        
        
    if 'Csv' in Export_Format:

        filename = 'Students_info.csv'
        
        with open(filename,'w',newline='') as file:
            csv_write = csv.writer(file)
            
            csv_write.writerow(New_Export_Heading)
            csv_write.writerows(New_Export_List)
        
        return send_file(filename,as_attachment=True)
            
    
    elif 'Excel' in Export_Format:
        
        filename = 'Export_Data.xlsx'
        Book = load_workbook('Export_File_Input_Format.xlsx')
        Sheet = Book.active

        for col_num, heading in enumerate(New_Export_Heading, 1):
            Sheet.cell(row=1, column=col_num, value=heading)
            
        for rows in New_Export_List:
            Sheet.append(rows)

        Book.save('Export_Data.xlsx')
        
        return send_file(filename,as_attachment=True)
    
    else:
        return jsonify({"res" : "Something Went Wrong!"})
    
        
        
@app.route('/log-out / <Pass>',methods = ['POST'])
def Log_out (Pass):
    
    if Pass == "bDr*^1t4t_@fj<lDda24Cz9*BM)I@u":
        
        Res = redirect('/admin-login')
        Res.delete_cookie('sat')
        return Res
    
    else:
        return "<h3>Access Denied !</h3>"

    

@app.route('/get-data-csv',methods=['GET','POST'])
@token_required
def Get_Csv_Data (admin):

    filters = request.get_json()
    
    
    if 'pass' in filters:
        if not filters['pass'] == "$2b$12$MAguwgtdDIGg3JAC9xKlcQC8EEinQsAfhtfs2Z7I8DAp9aG":
            return "Unauthorized!"
        
    else:
        return "Unauthorized!"
    
    filters = filters['data']
    College = filters[0]
    Course  = filters[1]
    Year_From  = filters[2]
    Year_To = filters[3]
    Payment  = filters[4]
    Mode = filters[5]
    Batch = filters[6]
    
    
    Mydb = mysql.connector.connect(
        host = "localhost",
        user = Mysql_Credentials.USER,
        password = Mysql_Credentials.PASS,
        database = 'sis'
        )
        
    cursor = Mydb.cursor()
    
    query = "SELECT * FROM students"
    First = True
    College_Trigger = False
    
    if not "All" in College :
        if First:
            query = f"{query} WHERE Inst_Key = %s"
            First = False
        else:
            query = f"{query} AND Inst_Key = %s"
            
        College_Trigger = True
            
    if not "All" in Course :
        if First:
            query = f"{query} WHERE Course_Name = '{Course}'"
            First = False
            
        else:
            query = f"{query} AND Course_Name = '{Course}'"
            
            
    if not "All" in Payment:
        if First:
            query = f"{query} WHERE  Payment_Status = '{Payment}'"
            First = False
        else:
            query = f"{query} AND Payment_Status = '{Payment}'"
            
    if not 'yyyy-MM-dd' in Year_From and not 'yyyy-MM-dd' in Year_To:
        if First:
            query = f"{query} WHERE DATE(Start_Date) BETWEEN '{Year_From}' AND '{Year_To}'"
            
        else:
            query = f"{query} AND DATE(Start_Date) BETWEEN '{Year_From}' AND '{Year_To}'"
              
    if not "All" in Mode:
        if First:
            query = f"{query} WHERE Mode = '{Mode}'"
            First = False
        else:
            query = f"{query} AND Mode = '{Mode}'"
            
    if not "All" in Batch:
        
        if First:
            query = f"{query} WHERE Batch = '{Batch}'"
            
        else:
            query = f"{query} AND Batch = '{Batch}'"
            
    query = f"{query};"

    if College_Trigger:
        cursor.execute(query,(College,))
    
    else:
        cursor.execute(query) 
         
    data = cursor.fetchall()
    
    if data:
        Students = []
        for Each_User in data:
            
            Name = Each_User[0]
            Last = Each_User[1]
            Phone = Each_User[2]
            Email = Each_User[3]
            Register_Number= Each_User[4]
            Institution_Name = Each_User[5]
            Mode = Each_User[6]
            Course_Name = Each_User[7]
            Total = Each_User[8]
            Entry_Date = Each_User[9]
            Payment_Status = Each_User[10]
            Inst_Key = Each_User[11]
            Password = Each_User[12]
            Certificate_Number = Each_User[13]
            End_Date = Each_User[14]
            Payment_Date = Each_User[15]
            Batch = Each_User[16]
            Start_Date = Each_User[17]
            Department = Each_User[18]
            
            Students.append(
                {
                    'First_Name' : Name.title(),
                    'Last_Name' : Last.title(),
                    'Phone' :  Phone,   
                    'Email' : Email,
                    'Register_Number' :  Register_Number,
                    'Institution_Name' : Institution_Name,
                    'Mode' : Mode,
                    'Course_Name' :  Course_Name,
                    'Total' :  Total,
                    'Entry_Date' : Entry_Date,
                    'Payment_Status' : Payment_Status.capitalize(),
                    'Inst_Key' : Inst_Key,
                    'Password' : Password,
                    'Certificate' : Certificate_Number,
                    'Start_Date' : Start_Date,
                    'End_Date' : End_Date,
                    'Payment_Date':Payment_Date,
                    'Department' : Department,
                    'Batch' : Batch
                }
                
            )
            
    else:
        Students = {'exist' : False}

    cursor.close()
    Mydb.close()
    
    return jsonify(Students)



@app.route('/admin-students')
@token_required
def Students_DashBoard(data):

    return render_template("Admin_Students_Page.html",IP=request.headers.get('X-Real-IP'))



@app.route('/import-file',methods=['POST'])
@token_required
def Import_File (admin):
    filename = request.files["File"]
    data = []
    
    if '.csv' in filename.filename:
        
        file_text = filename.read().decode('utf-8')
        Reader = csv.DictReader(file_text.splitlines())
        

        for each_user in Reader:
            data.append(each_user)
            
    elif '.xlsx' in filename.filename:
        
        Book = load_workbook(filename)
        Sheet = Book.active

        Rows = Sheet.rows

        Headers = [cell.value for cell in next(Rows)]

        for rows in Rows:
            Temp_data = {}
            
            for title,cell in zip(Headers,rows):
                Temp_data[title] = cell.value
                
            data.append(Temp_data)
            
        
    else:
        return redirect('/admin-students')
    
    
    if data:
        
        Mydb = mysql.connector.connect(
        host = "localhost",
        user = Mysql_Credentials.USER,
        password = Mysql_Credentials.PASS,
        database = 'sis'
        
        )
        
        cursor = Mydb.cursor()
        
        cursor.execute('SELECT * FROM Key_Dictionary')
        Key_List = cursor.fetchall()
        
        
        for each_user in data:
            
            if "Last_Name" in each_user:
                First_Name = each_user['Name'].lower()
                Last_Name = each_user['Last_Name'].lower()
                
            else:
                Name_Split = Clean_Data(each_user['Name']).Name_Split()
                
                First_Name = Name_Split[0].lower()
                Last_Name = Name_Split[1].lower()
                

            # datetime_object = datetime.strptime("", "%B %d %Y")
            # string = datetime_object.strftime("%Y-%m-%d %H:%M:%S")
                
            dt = datetime.now()
            Phone = Clean_Data(each_user.get("Phone","None")).Phone_Num_Clean()
            Email = each_user.get("Email","None")
            Register_Number= each_user.get("Usn","None")
            Institution_Name = each_user.get("College","None")
            Course_Name = each_user.get("Domain","None")
            Total = each_user.get("Total","None")
            Entry_Date = each_user.get("Entry_Date",str(dt.strftime("%Y-%m-%d")))
            Payment_Status = each_user.get("Payment_Status","None")
            Mode = each_user.get('Mode','None')
            Payment_Date = Entry_Date
            Department = each_user.get("Branch","None")
            
            # cursor.execute("SELECT First_Name FROM students WHERE Phone = %s AND Register_Number = %s AND Course_Name = %s;",(Phone,Register_Number,Course_Name,))
            # if_data_exist = cursor.fetchall()
            
            # if if_data_exist :
            #     pass
            
            if True:
                
                Key_Process = Inst_Process(Register_Number,Institution_Name,Key_List).Process()
        
                Inst_Key = Key_Process['inst_key']
                
                if not Key_Process['got']:
                    cursor.execute("INSERT INTO Key_Dictionary (Reg_key, Inst) VALUES (%s , %s)",Key_Process['keys'])
                    Key_List.append(Key_Process['keys'])
                    
                
                # Entry_Date = DateTimeProcess(Entry_Date).Get()
                # Payment_Date = DateTimeProcess(Payment_Date).Get()
                Password = Random_Password(10).Generate()
                Batch = Credentials.Batch
                
                cursor.execute("""INSERT INTO students (First_Name, Last_Name, Phone,
                    Email , Register_Number, Institution_Name, Mode,Course_Name,
                    Total, Entry_Date,Payment_Status,Inst_Key,Password,Payment_Date,Batch,Department) 
                    VALUES(IFNULL(%s,"Not defined"),IFNULL(%s,"Not defined"),IFNULL(%s,"Not defined"),IFNULL(%s,"Not defined"),IFNULL(%s,"Not defined"),IFNULL(%s,"Not defined"),IFNULL(%s,"Not defined"),IFNULL(%s,"Not defined"),IFNULL(%s,"Not defined"),IFNULL(%s,"Not defined"),IFNULL(%s,"Not defined"),IFNULL(%s,"Not defined"),IFNULL(%s,"Not defined"),IFNULL(%s,"Not defined"),IFNULL(%s,"Not defined"),IFNULL(%s,"Not defined"));"""
                    ,(First_Name,Last_Name,Phone,Email,Register_Number.upper(),Institution_Name,
                    Mode,Course_Name.lower(),Total,Entry_Date,Payment_Status,Inst_Key,Password,Payment_Date,Batch,Department))
        
        
        Mydb.commit()
        cursor.close()
        Mydb.close()
        
        return redirect('/admin-students')
    
    else:
        return redirect("/admin-students")
        
    


@app.route('/admin')
def Admin_Page ():

    return redirect('/admin-login')
    


@app.route('/admin-login')
def index ():
    token = request.cookies.get('sat')
    
    if token:
    
        try:
            jwt.decode(token, app.config['SECRET_KEY'],algorithms=['HS256'])
            return redirect('/admin-students')
            
        except:
            return render_template("Admin_Login.html")
        
    else:
        return render_template("Admin_Login.html")
    


@app.route('/login-data',methods = ['POST'])
def Login_process():
    
    Name_Email = request.form['Name_Email']
    Password = request.form['Password']
    
    if Name_Email and Password:
        
        Name_Email_found = False

        mydb = mysql.connector.connect(
            host = 'localhost',
            user = Mysql_Credentials.USER,
            password = Mysql_Credentials.PASS,
            database = 'sis'
        )
        
        cursor = mydb.cursor()
        
        if "@" in Name_Email:
            
            cursor.execute("SELECT Email, Password , First_Name FROM admin WHERE Email = %s",(Name_Email,))
            data = cursor.fetchall()
            cursor.close()
            mydb.close()
            
            if data:
                Hash_P = Hash_Password(Password)
                Email = data[0][0]
                
                if Email == Name_Email and Hash_P.Validate(data[0][1]):
        
                    Name_Email_found = True

                    token = jwt.encode({'user': request.headers.get('X-Real-IP'),'exp': datetime.utcnow() + timedelta(minutes=30)}, app.config['SECRET_KEY'])
                    
                    with open("Ips.txt",'a') as file:
                        file.write(f"Time :{datetime.now()}, Ip : {request.headers.get('X-Real-IP')}\n")
                    
                else:
                    flash("Incorrect Password !",'error')
                    return redirect('/admin-login')
            
            else:
              flash("Email does't exists!")
              return redirect('/admin-login')
              
        else:
            cursor.execute("SELECT Email, Password , First_Name FROM admin WHERE First_Name = %s",(Name_Email.lower(),))
            data = cursor.fetchall()
            cursor.close()
            mydb.close()
            
            if data:
                Hash_P = Hash_Password(Password)
                Name = data[0][2]
                
                if Name == Name_Email and Hash_P.Validate(data[0][1]):
        
                    Name_Email_found = True
                    
                    token = jwt.encode({'user': request.headers.get('X-Real-IP'),'exp': datetime.utcnow() + timedelta(minutes=30)}, app.config['SECRET_KEY'])
                    
                    with open("Ips.txt",'a') as file:
                        file.write(f"Time :{datetime.now()}, Ip : {request.headers.get('X-Real-IP')}\n")
                    
                else:
                    flash("Incorrect Password !",'error')
                    return redirect('/admin-login')
            
            else:
              flash("Name does't exists!")
              return redirect('/admin-login')
              
        if Name_Email_found:
            if session.get('redirect_url'):
                Res = session.get('redirect_url')
                session.pop('redirect_url')
                
            else:
                Res = '/admin-students'
                
            Res = redirect(Res)
            Res.set_cookie('sat',token,httponly=True)
            return Res
            
        else:
            return "Some thing wrong Happend!"
        
    else:
        flash("Invalid Input!",'error')
        return redirect('/admin-login')
       
       
if __name__ == "__main__":
    app.run(debug=True)
    